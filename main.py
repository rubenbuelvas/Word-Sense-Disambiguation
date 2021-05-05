import numpy as np
import pandas as pd
import csv
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import confusion_matrix, accuracy_score, f1_score
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.neural_network import MLPClassifier
from nltk.stem import LancasterStemmer
from nltk.tokenize import word_tokenize
from openpyxl import Workbook
from nltk.corpus import stopwords
from sklearn.model_selection import cross_validate

# Constants

lancaster_stemmer = LancasterStemmer()
CATEGORIES_STOPWORDS = ['/.', '/,', '[', ']', '/(', '/)', '\n', "/'", "/''", '/``', '/:']
CATEGORIES_STOPWORDS_EXTRA = ['/IN', '/DT', '/CC']
PUNCTUATION_STOPWORDS = ['.', ',', '[', ']', '(', ')', '\n', "'", "''", '``', ':', ';', '{', '}', '"']
ENGLISH_STOPWORDS = []
test_size = 0.3

# Config

DATASETS = ['gc', 'nw', 'ws']
MODELS = ['nb', 'dt', 'mlp']
CONFIG = {
    'stopwords': 'all',  # ['all', 'punctuation', 'none']
    'gc': {
        'filename': 'gc_dataset.csv',
        'n_words': 2,
        'vectorizer': 'count',  # ['count', 'tfidf']
        'mlp': {
            'solver': 'adam',  # ['lbfgs', 'sgd', 'adam']
            'hidden_layer_sizes': (100,)
        },
        'dt': {
            'depth': 100
        }
    },
    'nw': {
        'filename': 'nw_dataset.csv',
        'n_words': 4,
        'vectorizer': 'count',  # ['count', 'tfidf']
        'mlp': {
            'solver': 'adam',  # ['lbfgs', 'sgd', 'adam']
            'hidden_layer_sizes': (100,)
        },
        'dt': {
            'depth': 100
        }
    },
    'ws': {
        'filename': 'ws_dataset.csv',
        'vectorizer': 'count',  # ['count', 'tfidf']
        'mlp': {
            'solver': 'adam',  # ['lbfgs', 'sgd', 'adam']
            'hidden_layer_sizes': (100,)
        },
        'dt': {
            'depth': 100
        }
    },
    'custom': {

    }
}


# Utilities

def print_progress_bar(iteration, total, prefix='', suffix='', decimals=1, length=100, fill='█', printEnd="\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}|', end=printEnd)
    # Print New Line on Complete
    if iteration == total:
        print()


# Feature extraction

def generate_datasets():
    gc_extraction()
    nw_extraction()
    whole_sentence_extraction()


# Grammatical classification extraction

def gc_extraction(stopwords=True, extra_stopwords=False, custom_n_words=None):
    if custom_n_words is None:
        n_words = CONFIG['gc']['n_words']
    else:
        n_words = custom_n_words
    data = []
    with open('interest.acl94.txt') as file:
        lines = file.readlines()
    separator = lines[1]
    lines.remove(separator)
    for line in lines:
        tmp = []
        line = line.split(' ')
        try:
            line.remove('======================================')
        except:
            pass
        if stopwords:
            for i in range(len(line)):
                stopword_found = False
                for stopword in CATEGORIES_STOPWORDS:
                    if line[i].find(stopword) != -1:
                        stopword_found = True
                        break
                if not stopword_found and extra_stopwords:
                    for stopword in CATEGORIES_STOPWORDS_EXTRA:
                        if line[i].find(stopword) != -1:
                            stopword_found = True
                            break
                if not stopword_found:
                    tmp.append(line[i])
            line = tmp
        nulls = []
        for _ in range(n_words):
            nulls.append('/VOID')
        line = nulls + line + nulls
        target_word_found = False
        for i in range(len(line)):
            if line[i].find('interest_') == 0:
                category = 'C' + line[i][9:10]
                line = line[i - n_words:i + n_words + 1]
                target_word_found = True
                break
            elif line[i].find('interests_') == 0:
                category = 'C' + line[i][10:11]
                line = line[i - n_words:i + n_words + 1]
                line.pop(n_words)
                target_word_found = True
                break
        if target_word_found:
            line.pop(n_words)
            for i in range(len(line)):
                try:
                    line[i] = line[i].split('/')[1]
                except:
                    line[i] = 'VOID'
            line = ' '.join(line)
            line = line.replace('VOID', '')
            line = [category, line]
            data.append(line)
    with open(CONFIG['gc']['filename'], 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    print(CONFIG['gc']['filename'] + ' generated')


# Natural words extraction

def nw_extraction(punctuation_stopwords=False, custom_n_words=None):
    if custom_n_words is None:
        n_words = CONFIG['gc']['n_words']
    else:
        n_words = custom_n_words
    data = []
    with open('interest-original.txt') as file:
        lines = file.readlines()
    separator = lines[1]
    lines.remove(separator)
    for line in lines:
        line = word_tokenize(line)
        try:
            line.remove('======================================')
        except:
            pass
        for i in range(len(line)):
            line[i] = lancaster_stemmer.stem(line[i])
        if punctuation_stopwords:
            line = list(filter(lambda x: x not in PUNCTUATION_STOPWORDS, line))
        nulls = []
        for _ in range(n_words):
            nulls.append('VOID')
        line = nulls + line + nulls
        target_word_found = False
        for i in range(len(line)):
            if line[i].find('interest_') == 0:
                category = 'C' + line[i][9:10]
                line = line[i - n_words:i + n_words + 1]
                line.pop(n_words)
                target_word_found = True
                break
            elif line[i].find('interests_') == 0:
                category = 'C' + line[i][10:11]
                line = line[i - n_words:i + n_words + 1]
                line.pop(n_words)
                target_word_found = True
                break
            elif not line[i].isalpha():
                line[i] = 'NUM'
        if target_word_found:
            line = ' '.join(line)
            line = line.replace('VOID', '')
            line = [category, line]
            data.append(line)
    with open(CONFIG['nw']['filename'], 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    print(CONFIG['nw']['filename'] + ' generated')


# TF-IDF feature extraction

def whole_sentence_extraction(punctuation_stopwords=True):
    data = []
    with open('interest-original.txt') as file:
        lines = file.readlines()
    separator = lines[1]
    lines.remove(separator)
    for line in lines:
        line = line.replace('======================================', '')
        tk_line = word_tokenize(line)
        target_word_found = False
        if punctuation_stopwords:
            for stopword in PUNCTUATION_STOPWORDS:
                line = line.replace(stopword, '')
        for i in range(len(tk_line)):
            if tk_line[i].find('interest_') == 0:
                category = 'C' + tk_line[i][9:10]
                line = line.replace(tk_line[i], 'interest')
                target_word_found = True
                break
            elif tk_line[i].find('interests_') == 0:
                category = 'C' + tk_line[i][10:11]
                line = line.replace(tk_line[i], 'interests')
                target_word_found = True
                break
        if target_word_found:
            line = [category, line]
            data.append(line)
    with open(CONFIG['ws']['filename'], 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    print(CONFIG['ws']['filename'] + ' generated')


# TF-IDF + GC

def tfidf_gc_extraction():
    pass


# Running

def run_models(model=None, dataset=None):
    print('RUNNING MODELS...')
    if dataset is None:
        datasets = DATASETS
    else:
        datasets = [dataset]
    if model is None:
        models = MODELS
    else:
        models = [model]
    for current_dataset in datasets:
        data = pd.read_csv(CONFIG[dataset]['filename'], header=None)
        print('Dataset: ' + current_dataset)
        y = data[0].values
        X = data[1]
        if CONFIG[dataset]['vectorizer'] == 'count':
            vectorizer = CountVectorizer(stop_words='english')
        else:
            vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
        X = vectorizer.fit_transform(X)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=0)
        for current_model in models:
            if current_model == 'nb':
                acc, f1 = naive_bayes(X_train, X_test, y_train, y_test)
                print('NB Accuracy: ' + str('{:.4f}'.format(acc)))
            elif current_model == 'dt':
                acc, f1 = decision_tree(X_train, X_test, y_train, y_test, CONFIG[current_dataset]['dt']['depth'])
                print('DT Accuracy: ' + str('{:.4f}'.format(acc)))
            elif current_model == 'mlp':
                acc, f1 = multilayer_perceptron(
                    X_train, X_test, y_train, y_test, CONFIG[dataset]['mlp']['hidden_layer_sizes'],
                    solver=CONFIG[current_dataset]['mlp']['solver']
                )
                print('MLP Accuracy: ' + str('{:.4f}'.format(acc)))


# Testing

def test_n_words(n_words_range=(1, 10)):
    print('TESTING FOR NW...')
    for i in range(n_words_range[0], n_words_range[1] + 1):
        nw_extraction(custom_n_words=i)
        run_models(model='nb', dataset='nw')
    print('TESTING FOT GC...')
    for i in range(n_words_range[0], n_words_range[1] + 1):
        gc_extraction(custom_n_words=i)
        run_models(model='nb', dataset='gc')


def test_dt_config(depth_range=(1, 40)):
    wb = Workbook()
    ws = wb.active
    current_row = 2
    for current_depth in range(depth_range[0], depth_range[1] + 1):
        ws.cell(column=1, row=current_row, value=current_depth)
        current_row += 1
    current_column = 2
    for dataset in DATASETS:
        data = pd.read_csv(CONFIG[dataset]['filename'], header=None)
        y = data[0].values
        X = data[1]
        if CONFIG[dataset]['vectorizer'] == 'count':
            vectorizer = CountVectorizer(stop_words='english')
        else:
            vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
        X = vectorizer.fit_transform(X)
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=0)
        ws.cell(column=current_column, row=1, value=dataset)
        current_row = 2
        for current_depth in range(depth_range[0], depth_range[1] + 1):
            acc, f1 = decision_tree(X_train, X_test, y_train, y_test, current_depth)
            ws.cell(column=current_column, row=current_row, value=current_depth)
            ws.cell(column=current_column, row=current_row, value=acc)
            current_row += 1
        current_column += 1
    wb.save(filename='dt_depth_test.xlsx')
    print('dt_depth_test.xlsx generated')


def test_mlp_config(selected_dataset=None, hidden_layer_range=((10, 1), (150, 3))):
    if selected_dataset is None:
        datasets = DATASETS
    else:
        datasets = [selected_dataset]
    for dataset in datasets:
        wb = Workbook()
        ws = wb.active
        current_row = 2
        current_column = 2
        i = 0
        l = (
            len(range(hidden_layer_range[0][1], hidden_layer_range[1][1])) *
            len(range(hidden_layer_range[0][0], int(hidden_layer_range[1][0]) + 1)
        ))
        for current_net_depth in range(hidden_layer_range[0][1], hidden_layer_range[1][1] + 1):
            data = pd.read_csv(CONFIG[dataset]['filename'], header=None)
            y = data[0].values
            X = data[1]
            if CONFIG[dataset]['vectorizer'] == 'count':
                vectorizer = CountVectorizer(stop_words='english')
            else:
                vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
            X = vectorizer.fit_transform(X)
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=0)
            ws.cell(column=current_column, row=1, value=dataset + ' d=' + str(current_net_depth))
            current_row = 2
            for current_layer_size in range(hidden_layer_range[0][0], int(hidden_layer_range[1][0]) + 1):
                if current_layer_size % 10 == 0:
                    acc, f1 = multilayer_perceptron(
                        X_train, X_test, y_train, y_test,
                        hidden_layer_sizes=[current_layer_size, current_net_depth],
                        solver=CONFIG[dataset]['mlp']['solver']
                    )
                    ws.cell(column=current_column, row=current_row, value=current_layer_size)
                    ws.cell(column=current_column, row=current_row, value=acc)
                    ws.cell(column=1, row=current_row, value=current_layer_size)
                    current_row += 1
                i += 1
                print_progress_bar(i + 1, l, prefix='mlp_test_with_' + dataset + '.xlsx', suffix='Complete', length=50)
            current_column += 1
        wb.save(filename='mlp_test_with_' + dataset + '.xlsx')


# Naive Bayes

def naive_bayes(X_train, X_test, y_train, y_test):
    nb = MultinomialNB()
    nb.fit(X_train, y_train)
    y_pred = nb.predict(X_test)
    return accuracy_score(y_test, y_pred), f1_score(y_test, y_pred, average='macro')


# Decision Tree

def decision_tree(X_train, X_test, y_train, y_test, depth):
    dt = DecisionTreeClassifier(max_depth=depth)
    dt.fit(X_train, y_train)
    y_pred = dt.predict(X_test)
    return accuracy_score(y_test, y_pred), f1_score(y_test, y_pred, average='macro')


# MultiLayer Perceptron

def multilayer_perceptron(X_train, X_test, y_train, y_test, hidden_layer_sizes, solver):
    mlp = MLPClassifier(hidden_layer_sizes=hidden_layer_sizes, max_iter=2000, solver=solver)
    mlp.fit(X_train, y_train)
    y_pred = mlp.predict(X_test)
    return accuracy_score(y_test, y_pred), f1_score(y_test, y_pred, average='macro')


# Main routine

if __name__ == '__main__':
    generate_datasets()
    test_mlp_config()
